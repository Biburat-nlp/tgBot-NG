import os
import random
import openpyxl
import time
from telebot import types
from datetime import datetime, timedelta
from telegram import Bot, Update
from telegram import Bot, Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, CallbackContext, MessageHandler, CallbackQueryHandler, filters


TELEGRAM_API_TOKEN = 'token'

OIV_IDS = {
    'ГБУ Жилищник района Марьино города Москвы': 1,
    'Управа Марьино': 2,
    'ГБУ «Автомобильные дороги ЮВАО»': 3,
    'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 4,
    'Управа Выхино-Жулебино': 5,
    'ГБУ Жилищник Нижегородского района города Москвы': 6,
    'Управа Нижегородский': 7,
    'ГБУ Жилищник района Капотня города Москвы': 8,
    'Управа Капотня': 9,
    'ГБУ Жилищник района Кузьминки города Москвы': 10,
    'Управа Кузьминки': 11,
    'ГБУ Жилищник района Лефортово города Москвы': 12,
    'Управа Лефортово': 13,
    'ГБУ Жилищник района Люблино города Москвы': 14,
    'Управа Люблино': 15,
    'ГБУ Жилищник района Некрасовка города Москвы': 16,
    'Управа Некрасовка': 17,
    'ГБУ Жилищник района Печатники города Москвы': 18,
    'Управа Печатники': 19,
    'ГБУ Жилищник района Текстильщики города Москвы': 20,
    'Управа Текстильщики': 21,
    'ГБУ Жилищник Рязанского района города Москвы': 22,
    'Управа Рязанский': 23,
    'ГБУ Жилищник Южнопортового района города Москвы': 24,
    'Управа Южнопортовый': 25,
    'Жилищная инспекция по ЮВАО': 26,
    'Префектура Юго-Восточного округа': 27
}

CATEGORY_IDS = {
    'Парки/скверы': 101,
    'ДТ': 102,
    'МКД': 103,
    'ОДХ': 104
}

USER_OIV_MAP = {
    '1008683615': {
        'oiv_ids': [12, 13],
        'category_ids': [101, 102, 103, 104]
    },
    '309204640': {'oiv_ids': [14, 15, 24, 25], 'category_ids': [101,102,103,104]},  # Александр Геннадьевич
    '248001485': {'oiv_ids': [14, 15, 24, 25], 'category_ids': [101,102,103,104]},
    '949805580': {'oiv_ids': [10,11], 'category_ids': [103]},# Альберт Рашитович
    '342617808': {'oiv_ids': [3], 'category_ids': [101,102,103,104]}, # Юрий АВД
    '5771868721': {'oiv_ids': [12, 13], 'category_ids': [101,102,103,104]}, # Владислав Лефортово
    '444818192': {'oiv_ids': [16, 17], 'category_ids': [101,102,103,104]}, # Руслан Мегамозг
    '773634578': {'oiv_ids': [16, 17], 'category_ids': [101,102,103,104]}, #Ксения Витальевна
    '1859497322': {'oiv_ids': [16, 17], 'category_ids': [101,102,103,104]}, #Виктория Мочалова (Некрасовка)
    '2124882080': {'oiv_ids': [8, 9], 'category_ids': [101,102,103,104]}, #Вороненко Константин (Капонтня)
    '1104241617': {'oiv_ids': [6, 7], 'category_ids': [101,102,103,104]}, #Ромашкина Мария (Нижегородский)
    '748848833': {'oiv_ids': [18], 'category_ids': [103]}, #Ярослав Путенцов(Печатники)
    '1968513890': {'oiv_ids': [16, 17], 'category_ids': [101,102,103,104]}, #Алиева Татьяна(Некрасовка)
    '682768205': {'oiv_ids': [22, 23], 'category_ids': [101,102,103,104]}, #Бочкова Алёна Владиславовна (Рязанский Управа)
    '1159531079': {'oiv_ids': [3], 'category_ids': [101,102,103,104]}, #Екатерина Лазарева (АВД)
    '637773050': {'oiv_ids':[16, 17], 'category_ids': [101, 102, 103, 104]}, #Кальдинова Наталья Игоревна (Некрасовка)
    '1118256309': {'oiv_ids':[6, 7], 'category_ids': [101,102,103,104]}, #Будаева Анастасия Николаевна (Нижегородский Управа)
    '1747380648': {'oiv_ids':[1, 2], 'category_ids': [101,102,103,104]}, #Деминова Людмила Ивановна (Марьино)
    '879277421': {'oiv_ids':[21], 'category_ids': [101,102,103,104]}, #Симоненкова Елена Сергеевна (Управа Текстильщики)
    '498139669': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Косов Олег (Выхино Жулебино Управа)
    '1779374656': {'oiv_ids':[22], 'category_ids': [101,102,103,104]}, #Растанова Анастасия Александровна (Рязанский гбу)
    '214038510': {'oiv_ids':[20, 21], 'category_ids': [101,102,103,104]}, #Проценко Леонид Дмитриевич (Текстильщики управа)
    '259685801': {'oiv_ids':[21], 'category_ids': [101,102,103,104]}, #Павлова Анастасия Евгеньевна (Текстильщики управа)
    '1304016323': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Комова Екатерина Германовна(Выхино-Жулебино)
    '2111490716': {'oiv_ids':[12, 13], 'category_ids': [101,102,103,104]}, #Соломатина Елена (Лефортово ГБУ)
    '957128044': {'oiv_ids':[24, 25], 'category_ids': [101,102,103,104]}, #Малородова Евгения (Южнопортовый)
    '5681222962': {'oiv_ids':[10], 'category_ids': [103]}, #Сержантова Анна (Кузьминки ГБУ)
    '199143077': {'oiv_ids':[3], 'category_ids': [101,102,103,104]},  #Титов Сергей (АВД)
    '184804530': {'oiv_ids':[20], 'category_ids': [101,102,103,104]},  #Мацеевский Максим (Текстильщики ГБУ)
    '284643670': {'oiv_ids':[12, 13], 'category_ids': [101,102,103,104]},  #Екатерина Карпушина (Лефортово Зам главы Управы)
    '891811153': {'oiv_ids':[12], 'category_ids': [101,102,103,104]},  #Сергеева Екатерина (Лефортово глава ГБУ)
    '1787572712': {'oiv_ids':[20], 'category_ids': [101,102,103,104]},  #Ворожейкина Светлана (Текстильщики ГБУ)
    '5807151468': {'oiv_ids':[18, 19], 'category_ids': [102]},  #Амет Анна (Печатники дворы)
    '1726689160': {'oiv_ids':[14, 15], 'category_ids': [101,102,103,104]}, #Кто-то из Люблино
    '1461794477': {'oiv_ids':[3], 'category_ids': [101,102,103,104]}, #Землянский Максим Начальник СМЦ ГБУ
    '5227822667': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то из Кузьминок
    '777757256': {'oiv_ids':[18, 19], 'category_ids': [101,102,103,104]}, #Елизавета Печатники
    '6322616283': {'oiv_ids':[14, 15], 'category_ids': [101,102,103,104]}, #Пруцкова Александра Ильинична Люблино
    '737424515': {'oiv_ids':[14, 15], 'category_ids': [103]}, #Калушка Ольга Владимировна Люблино
    '916001760': {'oiv_ids':[14, 15], 'category_ids': [102,103]}, #Зайцева Полина Сергеевна Люблино
    '1104172214': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то из Кузьминок
    '990914503': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то из Кузьминок
    '5213710149': {'oiv_ids':[18, 19], 'category_ids': [101,102,103,104]}, #Кто-то Печатники
    '1208819533': {'oiv_ids':[3], 'category_ids': [101,102,103,104]}, #Кто-то из АВД
    '1065318763': {'oiv_ids':[10, 11], 'category_ids': [101,102,104]}, #Абрамов Никита Сергеевич Кузьминки
    '1479535368': {'oiv_ids':[6, 7], 'category_ids': [101,102,103,104]}, #Кто-то (Нижегородский)
    '1810562708': {'oiv_ids':[6, 7], 'category_ids': [101,102,103,104]}, #Кто-то (Нижегородский)
    '433713437': {'oiv_ids':[6, 7], 'category_ids': [101,102,103,104]}, #Киряков Александр Владиславович (Нижегородский)
    '399125424': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то из Кузьминок
    '1333864717': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то из Кузьминок
    '534595947': {'oiv_ids':[10, 11], 'category_ids': [103]}, #Кто-то из Кузьминок
    '5691778916': {'oiv_ids':[16, 17], 'category_ids': [101,102,103,104]}, #Зотов Сергей (Некрасовка. Зам.Главы Управы)
    '1958311512': {'oiv_ids':[14, 15], 'category_ids': [101,102,103,104]},  # Дурова Элина (Люблино Зам. Главы Управы)
    '1201993583': {'oiv_ids':[8, 9], 'category_ids': [101,102,103,104]}, #Шитиков Михаил (Капотня Зам. Главы Управы)
    '738501775': {'oiv_ids':[16, 17], 'category_ids': [101,102,103,104]}, #Хромова Елена (Некрасовка Глава Управы)
    '139355601': {'oiv_ids':[20, 21], 'category_ids': [101,102,103,104]}, #Кто-то Текстильщики
    '405402236': {'oiv_ids':[6, 7], 'category_ids': [101,102,103,104]}, #Хозяенок Игорь (Нижегородский Зам. Главы Управы)
    '266830829': {'oiv_ids':[24, 25], 'category_ids': [101,102,103,104]}, #Квачахия Рональд (Южнопортовый Зам. Главы Управы)
    '7383347962': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Дивин Александр (Выхино-Жулебино Руководитель ГБУ)
    '5220557356': {'oiv_ids':[18, 19], 'category_ids': [101,102,103,104]}, #Кузьмичёв Алексей (Печатники Зам. Главы Управы)
    '2140153164': {'oiv_ids':[18, 19], 'category_ids': [101,102,103,104]}, #Кто-то (Печатники)
    '541281446': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Кто-то (Выхино)
    '1159278532': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '458394303': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '1442199120': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '1615407360': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '1926075202': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '6778633208': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '5650866862': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]},  # Кто-то (Кузьминки)
    '420412441': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '221474889': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '5293442369': {'oiv_ids':[12, 13], 'category_ids': [101,102,103,104]}, #Кто-то (Лефортово)
    '7428321359': {'oiv_ids':[12, 13], 'category_ids': [101,102,103,104]}, #Начальник 8ого участка (Лефортово)
    '1677640950': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '755792631': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '6034511872': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '651322495': {'oiv_ids':[11], 'category_ids': [101,102,103,104]}, #Кто-то (управа Кузьминки)
    '6468976698': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '1530446114': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '5288775403': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '7529825867': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '5246328211': {'oiv_ids':[12, 13], 'category_ids': [101]}, #Почивалова Елена Викторовна (Лефортово)
    '1045446604': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Кто-то (Выхино-Жулебино)
    '5108733398': {'oiv_ids':[10, 11], 'category_ids': [101,102,103,104]}, #Кто-то (Кузьминки)
    '690416303': {'oiv_ids':[12, 13], 'category_ids': [101,102,103,104]}, #Кто-то (Лефортово)
    '1348896184': {'oiv_ids':[4, 5], 'category_ids': [101,102,103,104]}, #Кто-то (Выхино-Жулебино)
    '6698335587': {'oiv_ids':[12, 13], 'category_ids': [102,103]}, #начальник участка
    '6993367364': {'oiv_ids':[12, 13], 'category_ids': [102,103]}, #начальник участка
    '5226409299': {'oiv_ids':[12, 13], 'category_ids': [101, 102, 103, 104]}, #КТо то Лефортово
    '5082179002': {'oiv_ids':[6, 7], 'category_ids': [101, 102, 103, 104]}, #Кто-то Нижегородский
}

DOWNLOAD_PATH = 'C:\\Users\\gamag\\Downloads'

def create_main_keyboard():
    keyboard = [
        [KeyboardButton("Монитор 💻")],
        [KeyboardButton("Портал 🏢")],
        [KeyboardButton("Статистика 📊")],
        [KeyboardButton("Просроченные сообщения ⏰")],
        [KeyboardButton("Сообщения на завтра Монитор 💻")],
        [KeyboardButton("Сообщения на завтра Портал 🏢")],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

async def start(update: Update, context: CallbackContext):
    await update.message.reply_text("Добро пожаловать! Пожалуйста, воспользуйтесь одной из кнопок ниже:", reply_markup=create_main_keyboard())

def escape_markdown_v2(text):
    escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    for char in escape_chars:
        text = text.replace(char, '\\' + char)
    return text


def format_date(date):
    return date.strftime('%H:%M:%S %d.%m.%Y')


def process_excel_for_user(file_path, user_id, deadline_filter=None, monitor_flag=None):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    notifications = []
    today = datetime.now().date()
    now = datetime.now().replace(microsecond=0)

    valid_statuses = ["Нет ответа", "Готовится ответ", "На доработке"]
    user_oiv_data = USER_OIV_MAP.get(str(user_id), {})
    user_oiv_ids = user_oiv_data.get('oiv_ids', [])
    user_category_ids = user_oiv_data.get('category_ids', [])

    for row in sheet.iter_rows(min_row=2):
        status = row[20].value
        current_monitor_flag = row[33].value
        deadline1 = row[19].value
        deadline2 = row[47].value
        object_id = row[9].value

        if status not in valid_statuses or (monitor_flag is not None and current_monitor_flag != monitor_flag):
            continue

        oiv_name = row[18].value
        oiv_id = OIV_IDS.get(oiv_name)
        category_name = row[36].value
        category_id = CATEGORY_IDS.get(category_name)

        if oiv_id is None or oiv_id not in user_oiv_ids:
            continue

        if monitor_flag == 'Да' and (category_id is None or category_id not in user_category_ids):
            continue

        address = row[6].value
        message_id = row[1].value
        problem_topic = row[12].value

        if current_monitor_flag == 'Да' and status == "Готовится ответ":
            deadline = deadline1
        else:
            deadline = deadline2 if current_monitor_flag == 'Да' else deadline1

        if isinstance(deadline, datetime):
            deadline = deadline.replace(microsecond=0)

            print(f"Checking deadline: {deadline}, now: {now}")

            if deadline < now:
                print(f"Skipping due to past deadline: {deadline}")
                continue

            if deadline_filter:
                if deadline.date() == today:
                    message = (f"Срок сегодня до {format_date(deadline)}\n"
                               f"Адрес: {address}\n"
                               f"Тема: {problem_topic}\n"
                               f"Номер сообщения: {message_id}\n"
                               f"Категория: {category_name}\n"
                               f"Ответственный ОИВ: {oiv_name}\n"
                               f"https://gorod.mos.ru/objects/{object_id}/messages#{message_id}")
                    escaped_message = escape_markdown_v2(message)
                    notifications.append(escaped_message)
            else:
                if now <= deadline <= now + timedelta(hours=2):
                    message = (f"До просрока менее двух часов! {format_date(deadline)}\n"
                               f"Адрес: {address}\n"
                               f"Тема: {problem_topic}\n"
                               f"Номер сообщения: {message_id}\n"
                               f"Категория: {category_name}\n"
                               f"Ответственный ОИВ: {oiv_name}\n"
                               f"https://gorod.mos.ru/objects/{object_id}/messages#{message_id}")
                    escaped_message = escape_markdown_v2(message)
                    notifications.append(escaped_message)

    return notifications


def count_messages_today(file_path, user_id):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    monitor_count = 0
    portal_count = 0
    total_count = 0

    today = datetime.now().date()
    now = datetime.now()

    valid_statuses = ["Нет ответа", "Готовится ответ", "На доработке"]
    user_oiv_data = USER_OIV_MAP.get(str(user_id), {})
    user_oiv_ids = user_oiv_data.get('oiv_ids', [])
    user_category_ids = user_oiv_data.get('category_ids', [])

    for row in sheet.iter_rows(min_row=2):
        status = row[20].value
        current_monitor_flag = row[33].value
        deadline1 = row[19].value
        deadline2 = row[47].value
        overdue_monitor = row[45].value

        if status not in valid_statuses:
            continue

        oiv_name = row[18].value
        oiv_id = OIV_IDS.get(oiv_name)
        category_name = row[36].value
        category_id = CATEGORY_IDS.get(category_name)

        if oiv_id is None or oiv_id not in user_oiv_ids:
            continue

        if current_monitor_flag == 'Да' and (category_id is None or category_id not in user_category_ids):
            continue

        if current_monitor_flag == 'Да' and status == "Готовится ответ":
            deadline = deadline1
        else:
            deadline = deadline2 if current_monitor_flag == 'Да' else deadline1

        if isinstance(deadline, datetime) and deadline.date() == today and deadline.time() >= now.time():
            if overdue_monitor == 'Да':
                continue
            total_count += 1
            if current_monitor_flag == 'Да':
                monitor_count += 1
            elif current_monitor_flag == 'Нет' or not current_monitor_flag:
                portal_count += 1

    return monitor_count, portal_count, total_count


def process_excel_for_user_tomorrow(file_path, user_id, monitor_flag=None):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    notifications = []
    tomorrow = datetime.now().date() + timedelta(days=1)

    valid_statuses = ["Нет ответа", "Готовится ответ", "На доработке"]
    user_oiv_data = USER_OIV_MAP.get(str(user_id), {})
    user_oiv_ids = user_oiv_data.get('oiv_ids', [])
    user_category_ids = user_oiv_data.get('category_ids', [])

    for row in sheet.iter_rows(min_row=2):
        status = row[20].value
        current_monitor_flag = row[33].value
        deadline1 = row[19].value
        deadline2 = row[47].value
        object_id = row[9].value

        if status not in valid_statuses or (monitor_flag is not None and current_monitor_flag != monitor_flag):
            continue

        oiv_name = row[18].value
        oiv_id = OIV_IDS.get(oiv_name)
        category_name = row[36].value
        category_id = CATEGORY_IDS.get(category_name)

        if oiv_id is None or oiv_id not in user_oiv_ids:
            continue

        if current_monitor_flag == 'Да' and (category_id is None or category_id not in user_category_ids):
            continue

        address = row[6].value
        message_id = row[1].value
        problem_topic = row[12].value

        if current_monitor_flag == 'Да' and status == "Готовится ответ":
            deadline = deadline1
        else:
            deadline = deadline2 if current_monitor_flag == 'Да' else deadline1

        if isinstance(deadline, datetime):
            deadline = deadline.replace(microsecond=0)

            if deadline.date() == tomorrow:
                message = (f"Срок завтра - {format_date(deadline)}\n"
                           f"Адрес: {address}\n"
                           f"Тема: {problem_topic}\n"
                           f"Номер сообщения: {message_id}\n"
                           f"Ответственный ОИВ: {oiv_name}, {category_name}\n"
                           f"https://gorod.mos.ru/objects/{object_id}/messages#{message_id}")
                escaped_message = escape_markdown_v2(message)
                notifications.append(escaped_message)
    return notifications


def get_random_waiting_message(file_path='waiting_messages.txt'):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            if lines:
                return random.choice(lines).strip()
            else:
                return "Бот думает, подождите немного..."
    except FileNotFoundError:
        return "Файл с сообщениями для ожидания не найден."


def get_prosrok_notifications(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    prosrok_notifications = {}
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    valid_statuses = ["Нет ответа", "Готовится ответ", "На доработке"]

    for row in sheet.iter_rows(min_row=2):
        prosrok_monitor = row[41].value
        current_monitor_flag = row[33].value
        status = row[20].value
        deadline1 = row[19].value
        deadline2 = row[47].value
        oiv_name = row[18].value
        message_id = row[1].value
        address = row[6].value
        problem_topic = row[12].value
        object_id = row[9].value
        category_name = row[36].value
        category_id = CATEGORY_IDS.get(category_name)

        if prosrok_monitor == 'Да' and status in valid_statuses:
            if current_monitor_flag == 'Да' and status == "Готовится ответ":
                deadline = deadline1
            else:
                deadline = deadline2 if current_monitor_flag == 'Да' else deadline1

            if isinstance(deadline, datetime) and (deadline.date() == today or deadline.date() == yesterday):
                message = (f"Просрочено - Дедлайн {format_date(deadline)}\n"
                           f"Адрес: {address}\n"
                           f"Тема: {problem_topic}\n"
                           f"Номер сообщения: {message_id}\n"
                           f"Ответственный ОИВ: {oiv_name}\n"
                           f"https://gorod.mos.ru/objects/{object_id}/messages#{message_id}")

                escaped_message = escape_markdown_v2(message)

                for user_id, oiv_list in USER_OIV_MAP.items():
                    if OIV_IDS.get(oiv_name) in oiv_list['oiv_ids'] and category_id in oiv_list['category_ids']:
                        if user_id not in prosrok_notifications:
                            prosrok_notifications[user_id] = []
                        prosrok_notifications[user_id].append(escaped_message)

    return prosrok_notifications


async def notify_user(user_id, messages):
    bot = Bot(token=TELEGRAM_API_TOKEN)
    for message in messages:
        try:
            await bot.send_message(chat_id=user_id, text=message, parse_mode='MarkdownV2')
            print(f"Message sent to {user_id}: {message}")
        except Exception as e:
            print(f"Failed to send message to {user_id}: {e}")


last_request_times = {}

# Минимальный интервал между запросами (в секундах)
REQUEST_INTERVAL = 10


async def check_request_interval(user_id):
    current_time = time.time()
    last_request_time = last_request_times.get(user_id, 0)

    if current_time - last_request_time < REQUEST_INTERVAL:
        return False
    else:
        last_request_times[user_id] = current_time
        return True


async def monitor(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)
    if latest_file:
        notifications = process_excel_for_user(latest_file, user_id, deadline_filter=True, monitor_flag='Да')
        count = len(notifications)
        if notifications:
            await notify_user(user_id, notifications)
        await update.message.reply_text(f"Сообщений на сегодня: {count}")
    else:
        await update.message.reply_text("Файл не найден!")


async def portal(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)
    if latest_file:
        notifications = process_excel_for_user(latest_file, user_id, deadline_filter=True, monitor_flag='Нет')
        count = len(notifications)
        if notifications:
            await notify_user(user_id, notifications)
        await update.message.reply_text(f"Сообщений на сегодня: {count}")
    else:
        await update.message.reply_text("Файл не найден!")


async def stats(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)
    if latest_file:
        monitor_count, portal_count, total_count = count_messages_today(latest_file, user_id)
        response = (f"Всего сообщений в работе со сроком сегодня: {total_count}\n"
                    f"Монитор: {monitor_count}\n"
                    f"Портал: {portal_count}")
        await update.message.reply_text(response)
    else:
        await update.message.reply_text("Файл не найден!")


async def prosrok_command(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)

    if latest_file:
        prosrok_notifications = get_prosrok_notifications(latest_file)
        count = 0

        if str(user_id) in prosrok_notifications:
            for message in prosrok_notifications[str(user_id)]:
                await context.bot.send_message(chat_id=user_id, text=message, parse_mode='MarkdownV2')
                count += 1

            await update.message.reply_text(f"Просроченных сообщений за вчера и сегодня: {count}")
        else:
            await update.message.reply_text("Нет просроченных сообщений за вчера и сегодня.")
    else:
        await update.message.reply_text("Файл не найден.")


async def portal_tomorrow(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)
    if latest_file:
        notifications = process_excel_for_user_tomorrow(latest_file, user_id, monitor_flag='Нет')
        count = len(notifications)

        if notifications:
            await notify_user(user_id, notifications)

        await update.message.reply_text(f"Сообщений на завтра (портал): {count}")
    else:
        await update.message.reply_text("Файл не найден!")


async def monitor_tomorrow(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id

    if not await check_request_interval(user_id):
        await update.message.reply_text("Пожалуйста, подождите 10 секунд перед следующим запросом. Бот тоже может уставать)")
        return

    random_waiting_message = get_random_waiting_message()
    await update.message.reply_text(random_waiting_message)

    latest_file = get_latest_downloaded_file(DOWNLOAD_PATH)
    if latest_file:
        notifications = process_excel_for_user_tomorrow(latest_file, user_id, monitor_flag='Да')
        count = len(notifications)

        if notifications:
            await notify_user(user_id, notifications)

        await update.message.reply_text(f"Сообщений на завтра (монитор): {count}")
    else:
        await update.message.reply_text("Файл не найден!")


def get_latest_downloaded_file(download_path):
    files = os.listdir(download_path)
    files = [os.path.join(download_path, f) for f in files if os.path.isfile(os.path.join(download_path, f))]
    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return files[0] if files else None


def main():
    application = Application.builder().token(TELEGRAM_API_TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Монитор 💻"), monitor))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Портал 🏢"), portal))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Статистика 📊"), stats))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Просроченные сообщения ⏰"), prosrok_command))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Сообщения на завтра Монитор"),monitor_tomorrow))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND & filters.Regex("Сообщения на завтра Портал"), portal_tomorrow))
    application.add_handler(CommandHandler("monitor", monitor))
    application.add_handler(CommandHandler("portal", portal))
    application.add_handler(CommandHandler("stats", stats))
    application.add_handler(CommandHandler("prosrok", prosrok_command))
    application.add_handler(CommandHandler("monitor_tomorrow", monitor_tomorrow))
    application.add_handler(CommandHandler("portal_tomorrow", portal_tomorrow))

    application.run_polling()


if __name__ == "__main__":
    main()